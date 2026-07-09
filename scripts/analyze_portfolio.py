#!/usr/bin/env python3
"""Weekly portfolio analysis, run by the Claude Routine.

Reads data/portfolio.xlsx (kept up to date by the weekly-fetch GitHub Action)
and the "History" sheet's weekly snapshots, then writes a markdown report to
reports/<date>.md covering performance, allocation, dividend changes, and
rule-based rebalance flags. Does not touch the network -- all data already
lives in the spreadsheet by the time this runs.
"""
import datetime
from collections import defaultdict
from pathlib import Path

import openpyxl

PORTFOLIO_PATH = Path(__file__).parent.parent / "data" / "portfolio.xlsx"
REPORTS_DIR = Path(__file__).parent.parent / "reports"

CONCENTRATION_THRESHOLD = 0.40
LOSS_THRESHOLD = -0.15
DIVIDEND_YIELD_CHANGE_THRESHOLD = 0.005  # 50 bps

MARKET_CURRENCY = {"AEX": "EUR", "NYSE": "USD", "NASDAQ": "USD"}


def find_header_row(ws, col, label, start_row=1):
    for row in range(start_row, ws.max_row + 1):
        if ws.cell(row=row, column=col).value == label:
            return row
    return None


def read_table_rows(ws, header_row, ticker_col=2):
    rows = []
    r = header_row + 1
    while True:
        value = ws.cell(row=r, column=ticker_col).value
        if not value or not isinstance(value, str) or ":" not in value:
            break
        rows.append(r)
        r += 1
    return rows


def load_holdings(ws):
    header_row = find_header_row(ws, col=2, label="Ticker", start_row=1)
    holdings = []
    for row in read_table_rows(ws, header_row):
        ticker = ws.cell(row=row, column=2).value
        naam = ws.cell(row=row, column=3).value
        aantal = ws.cell(row=row, column=4).value or 0
        aankoopprijs = ws.cell(row=row, column=5).value or 0
        huidige_waarde = ws.cell(row=row, column=6).value or 0
        dividend_pct = ws.cell(row=row, column=9).value or 0
        market = ws.cell(row=row, column=10).value or ""
        total_value = aantal * huidige_waarde
        cost_basis = aantal * aankoopprijs
        pct_change = (huidige_waarde - aankoopprijs) / aankoopprijs if aankoopprijs else 0
        holdings.append({
            "ticker": ticker, "naam": naam, "aantal": aantal,
            "aankoopprijs": aankoopprijs, "huidige_waarde": huidige_waarde,
            "dividend_pct": dividend_pct, "market": market,
            "total_value": total_value, "cost_basis": cost_basis,
            "pct_change": pct_change,
        })
    return holdings


def load_history(wb):
    if "History" not in wb.sheetnames:
        return {}
    ws = wb["History"]
    by_date = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        date, ticker, naam, aantal, price, currency, dividend_yield, total_value = row
        by_date[str(date)][ticker] = {
            "price": price, "dividend_yield": dividend_yield, "total_value": total_value,
        }
    return by_date


def build_report(holdings, history_by_date):
    lines = []
    today = datetime.date.today().isoformat()
    lines.append(f"# Weekly Portfolio Report - {today}")
    lines.append("")

    portfolio_value = sum(h["total_value"] for h in holdings)
    portfolio_cost = sum(h["cost_basis"] for h in holdings)
    portfolio_gain = portfolio_value - portfolio_cost
    portfolio_gain_pct = (portfolio_gain / portfolio_cost) if portfolio_cost else 0
    weighted_yield = (
        sum(h["total_value"] * h["dividend_pct"] for h in holdings) / portfolio_value
        if portfolio_value else 0
    )

    lines.append("## Performance Summary")
    lines.append("")
    lines.append(f"- **Total value:** {portfolio_value:,.2f} (mixed currencies, see caveat below)")
    lines.append(f"- **Total cost basis:** {portfolio_cost:,.2f}")
    lines.append(f"- **Unrealized gain/loss:** {portfolio_gain:,.2f} ({portfolio_gain_pct:+.2%})")
    lines.append(f"- **Weighted-average dividend yield:** {weighted_yield:.2%}")
    lines.append("")
    lines.append("| Ticker | Naam | Shares | Cost Basis | Price | Value | Gain/Loss | Div Yield |")
    lines.append("|---|---|---|---|---|---|---|---|")
    for h in sorted(holdings, key=lambda x: -x["total_value"]):
        lines.append(
            f"| {h['ticker']} | {h['naam']} | {h['aantal']} | {h['aankoopprijs']:.2f} | "
            f"{h['huidige_waarde']:.2f} | {h['total_value']:.2f} | {h['pct_change']:+.2%} | "
            f"{h['dividend_pct']:.2%} |"
        )
    lines.append("")

    lines.append("## Allocation")
    lines.append("")
    by_market = defaultdict(float)
    for h in holdings:
        by_market[h["market"]] += h["total_value"]
    lines.append("| Ticker | Naam | % of Portfolio |")
    lines.append("|---|---|---|")
    for h in sorted(holdings, key=lambda x: -x["total_value"]):
        pct = h["total_value"] / portfolio_value if portfolio_value else 0
        lines.append(f"| {h['ticker']} | {h['naam']} | {pct:.1%} |")
    lines.append("")
    lines.append("**By market:** " + ", ".join(
        f"{m}: {v / portfolio_value:.1%}" for m, v in by_market.items()
    ))
    lines.append("")

    dates = sorted(history_by_date.keys())
    latest_date = dates[-1] if dates else None
    prev_date = dates[-2] if len(dates) > 1 else None
    dividend_decreases = []
    if latest_date and prev_date:
        latest = history_by_date[latest_date]
        prev = history_by_date[prev_date]
        lines.append(f"## Dividend Changes ({prev_date} -> {latest_date})")
        lines.append("")
        dividend_change_lines = []
        for h in holdings:
            t = h["ticker"]
            if t in latest and t in prev and prev[t]["dividend_yield"] is not None:
                delta = (latest[t]["dividend_yield"] or 0) - (prev[t]["dividend_yield"] or 0)
                if abs(delta) >= DIVIDEND_YIELD_CHANGE_THRESHOLD:
                    direction = "increased" if delta > 0 else "decreased"
                    dividend_change_lines.append(
                        f"- **{h['naam']} ({t})**: trailing dividend yield {direction} from "
                        f"{prev[t]['dividend_yield']:.2%} to {latest[t]['dividend_yield']:.2%}"
                    )
                    if delta < 0:
                        dividend_decreases.append(h)
        lines.extend(dividend_change_lines if dividend_change_lines else ["No material dividend yield changes this week."])
        lines.append("")
    else:
        lines.append("## Dividend Changes")
        lines.append("")
        lines.append("Not enough history yet for a week-over-week comparison (need at least two weekly snapshots).")
        lines.append("")

    lines.append("## Flags & Suggestions")
    lines.append("")
    flags = []
    for h in holdings:
        alloc = h["total_value"] / portfolio_value if portfolio_value else 0
        if alloc >= CONCENTRATION_THRESHOLD:
            flags.append(
                f"- **Concentration risk:** {h['naam']} ({h['ticker']}) is {alloc:.0%} of the "
                f"portfolio -- consider trimming to reduce single-name risk."
            )
        if h["pct_change"] <= LOSS_THRESHOLD:
            flags.append(
                f"- **Underwater position:** {h['naam']} ({h['ticker']}) is down "
                f"{h['pct_change']:.1%} from cost basis -- worth reviewing the thesis."
            )
    for h in dividend_decreases:
        flags.append(
            f"- **Possible dividend cut:** {h['naam']} ({h['ticker']})'s trailing yield dropped "
            f"week over week -- check for a dividend cut announcement."
        )

    currencies = {MARKET_CURRENCY.get(h["market"], "USD") for h in holdings}
    if len(currencies) > 1:
        flags.append(
            f"- **Data caveat:** holdings span multiple currencies ({', '.join(sorted(currencies))}) "
            f"but the portfolio total sums raw numbers with no FX conversion -- treat the total value "
            f"as directional, not exact."
        )
    if len({h["market"] for h in holdings}) <= 2 and len(holdings) <= 5:
        flags.append(
            "- **Diversification:** only a handful of holdings across 1-2 markets -- "
            "concentration risk from a small position count, independent of any single name's weight."
        )

    lines.extend(flags if flags else ["No flags this week."])
    lines.append("")

    return "\n".join(lines)


def main():
    wb = openpyxl.load_workbook(PORTFOLIO_PATH, data_only=False)
    ws = wb["Sheet1"]
    holdings = load_holdings(ws)
    history_by_date = load_history(wb)

    report = build_report(holdings, history_by_date)

    REPORTS_DIR.mkdir(exist_ok=True)
    out_path = REPORTS_DIR / f"{datetime.date.today().isoformat()}.md"
    out_path.write_text(report)

    print(report)
    print(f"\nReport written to {out_path}")


if __name__ == "__main__":
    main()
