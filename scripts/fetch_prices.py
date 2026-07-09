#!/usr/bin/env python3
"""Weekly market-data updater for data/portfolio.xlsx.

Runs on GitHub Actions (which has normal internet access) via the
"weekly-fetch" workflow. Pulls current prices and dividend history from
Yahoo Finance (yfinance) and:

  1. Updates "Huidige waarde" (current price) and "Dividend %" per holding.
  2. Updates the current quarter's cell in the Dividends table with the
     actual cash dividends received this quarter (per-share dividend x shares).
  3. Appends a snapshot row to the "History" sheet so analyze_portfolio.py
     can compare week over week.

Formulas already in the sheet (% Verandering, Totale waarde, dividend
totals) are left untouched -- Excel recalculates them on open.
"""
import datetime
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import yfinance as yf

sys.path.insert(0, str(Path(__file__).parent))
from ticker_map import to_yahoo_symbol

PORTFOLIO_PATH = Path(__file__).parent.parent / "data" / "portfolio.xlsx"

CURRENCY_BY_SUFFIX = {
    ".AS": "EUR",
    ".PA": "EUR",
    ".DE": "EUR",
    ".L": "GBP",
}


def infer_currency(symbol: str) -> str:
    for suffix, currency in CURRENCY_BY_SUFFIX.items():
        if symbol.endswith(suffix):
            return currency
    return "USD"


def find_header_row(ws, col, label, start_row=1):
    for row in range(start_row, ws.max_row + 1):
        if ws.cell(row=row, column=col).value == label:
            return row
    return None


def read_table_rows(ws, header_row, ticker_col=2):
    rows = []
    r = header_row + 1
    while True:
        value = ws.cell(row=r, column=ticker_col).value
        if not value or not isinstance(value, str) or ":" not in value:
            break
        rows.append(r)
        r += 1
    return rows


def fetch_ticker_data(symbol: str):
    tk = yf.Ticker(symbol)
    hist = tk.history(period="5d")
    if hist.empty:
        raise ValueError(f"no price history returned for {symbol}")
    price = float(hist["Close"].dropna().iloc[-1])

    try:
        divs = tk.dividends
    except Exception:
        divs = pd.Series(dtype=float)

    return price, divs


def main():
    issues = []
    wb = openpyxl.load_workbook(PORTFOLIO_PATH)
    ws = wb["Sheet1"]

    holdings_header = find_header_row(ws, col=2, label="Ticker", start_row=1)
    dividends_header = find_header_row(ws, col=2, label="Ticker", start_row=holdings_header + 1)
    if holdings_header is None or dividends_header is None:
        raise RuntimeError("Could not locate both header rows in Sheet1")

    holdings_rows = read_table_rows(ws, holdings_header)
    dividend_rows = read_table_rows(ws, dividends_header)
    dividend_row_by_ticker = {ws.cell(row=r, column=2).value: r for r in dividend_rows}

    history_ws = wb["History"]
    today = datetime.date.today().isoformat()
    quarter = (datetime.date.today().month - 1) // 3 + 1
    quarter_col = {1: 5, 2: 6, 3: 7, 4: 8}[quarter]  # E, F, G, H
    now_year = datetime.date.today().year

    for row in holdings_rows:
        ticker = ws.cell(row=row, column=2).value
        naam = ws.cell(row=row, column=3).value
        aantal = ws.cell(row=row, column=4).value
        market = ws.cell(row=row, column=10).value
        symbol = to_yahoo_symbol(ticker, market)

        try:
            price, divs = fetch_ticker_data(symbol)
        except Exception as exc:
            issues.append(f"{ticker} ({symbol}): failed to fetch price - {exc}")
            continue

        currency = infer_currency(symbol)
        ws.cell(row=row, column=6, value=round(price, 2))  # Huidige waarde

        ttm_cutoff = pd.Timestamp.now(tz=divs.index.tz) - pd.Timedelta(days=365) if not divs.empty else None
        ttm_divs = divs[divs.index >= ttm_cutoff] if ttm_cutoff is not None else divs
        ttm_per_share = float(ttm_divs.sum()) if not ttm_divs.empty else 0.0
        dividend_yield = (ttm_per_share / price) if price else 0.0
        ws.cell(row=row, column=9, value=round(dividend_yield, 4))  # Dividend %

        div_row = dividend_row_by_ticker.get(ticker)
        if div_row and not divs.empty:
            q_divs = divs[(divs.index.year == now_year) & (((divs.index.month - 1) // 3 + 1) == quarter)]
            q_total_per_share = float(q_divs.sum())
            q_total_cash = round(q_total_per_share * (aantal or 0), 2)
            ws.cell(row=div_row, column=quarter_col, value=q_total_cash)

        history_ws.append([
            today, ticker, naam, aantal, round(price, 2), currency,
            round(dividend_yield, 4), round(price * (aantal or 0), 2),
        ])

    wb.save(PORTFOLIO_PATH)

    if issues:
        print("Fetch completed with issues:")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print("Fetch completed successfully for all holdings.")


if __name__ == "__main__":
    main()
